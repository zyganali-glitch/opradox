from io import BytesIO
from typing import Any, Dict
import pandas as pd
from fastapi import HTTPException
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule


def run(df: pd.DataFrame, params: Dict[str, Any]) -> Dict[str, Any]:
    # Parametre kontrolü
    value_column = params.get("value_column")
    lower_threshold = params.get("lower_threshold")
    upper_threshold = params.get("upper_threshold")
    lower_color = params.get("lower_color", "FFFF0000")  # default kırmızı (ARGB)
    upper_color = params.get("upper_color", "FF00FF00")  # default yeşil (ARGB)

    # value_column boşsa ilk numeric auto-seç
    if not value_column:
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        if numeric_cols:
            value_column = numeric_cols[0]
        else:
            raise HTTPException(status_code=400, detail="Sayısal sütun bulunamadı. Lütfen value_column belirtin.")
    
    if lower_threshold is None and upper_threshold is None:
        # En az bir eşik gerekli - ya da default olarak Q1/Q3 kullan
        numeric_series = pd.to_numeric(df[value_column], errors="coerce").dropna()
        if len(numeric_series) > 0:
            lower_threshold = numeric_series.quantile(0.25)
            upper_threshold = numeric_series.quantile(0.75)

    if value_column not in df.columns:
        raise HTTPException(
            status_code=400,
            detail=f"{value_column} sütunu bulunamadı, mevcut sütunlar: {list(df.columns)[:10]}"
        )

    # Sayısal dönüşüm
    df[value_column] = pd.to_numeric(df[value_column], errors="coerce")

    # Excel dosyası oluşturma
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        ws = writer.sheets["Data"]

        col_idx = None
        for idx, col in enumerate(df.columns, start=1):
            if col == value_column:
                col_idx = idx
                break

        if col_idx is None:
            raise HTTPException(status_code=400, detail="value_column sütunu Excel'de bulunamadı")

        # Koşullu biçimlendirme kuralları
        if lower_threshold is not None:
            try:
                low_val = float(lower_threshold)
            except Exception:
                raise HTTPException(status_code=400, detail="lower_threshold sayısal olmalı")
            red_fill = PatternFill(start_color=lower_color, end_color=lower_color, fill_type="solid")
            ws.conditional_formatting.add(
                f"{ws.cell(row=2, column=col_idx).coordinate}:{ws.cell(row=ws.max_row, column=col_idx).coordinate}",
                CellIsRule(operator="lessThan", formula=[str(low_val)], fill=red_fill)
            )

        if upper_threshold is not None:
            try:
                up_val = float(upper_threshold)
            except Exception:
                raise HTTPException(status_code=400, detail="upper_threshold sayısal olmalı")
            green_fill = PatternFill(start_color=upper_color, end_color=upper_color, fill_type="solid")
            ws.conditional_formatting.add(
                f"{ws.cell(row=2, column=col_idx).coordinate}:{ws.cell(row=ws.max_row, column=col_idx).coordinate}",
                CellIsRule(operator="greaterThan", formula=[str(up_val)], fill=green_fill)
            )

    output.seek(0)

    summary = {
        "value_column": value_column,
        "lower_threshold": lower_threshold,
        "upper_threshold": upper_threshold,
        "rows": len(df),
        "colored_cells": {
            "below_lower": int((df[value_column] < float(lower_threshold)).sum()) if lower_threshold is not None else 0,
            "above_upper": int((df[value_column] > float(upper_threshold)).sum()) if upper_threshold is not None else 0,
        }
    }
    
    # Python kod özeti
    technical_details = {
        "scenario": "highlight_values_by_threshold",
        "parameters": {"value_column": value_column, "lower_threshold": lower_threshold, "upper_threshold": upper_threshold},
        "stats": summary["colored_cells"],
        "python_code": f"""```python
import pandas as pd

df = pd.read_excel('dosya.xlsx')

# Eşik değerlerine göre renklendirme (Excel'de yapılır)
lower = {lower_threshold}
upper = {upper_threshold}

below = (df['{value_column}'] < lower).sum() if lower else 0
above = (df['{value_column}'] > upper).sum() if upper else 0

print(f"Alt sınırın altında: {{below}}, Üst sınırın üstünde: {{above}}")
```"""
    }

    return {
        "summary": summary,
        "technical_details": technical_details,
        "df_out": df,
        "excel_bytes": output,
        "excel_filename": "highlighted_values.xlsx"
    }