from io import BytesIO
from typing import Any, Dict
import pandas as pd
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows


def run(df: pd.DataFrame, params: Dict[str, Any]) -> Dict[str, Any]:
    # Parametre kontrolü
    value_column = params.get("value_column")
    if not value_column:
        raise HTTPException(status_code=400, detail="value_column parametresi gerekli")
    if value_column not in df.columns:
        raise HTTPException(
            status_code=400,
            detail=f"value_column '{value_column}' bulunamadı, mevcut sütunlar: {list(df.columns)}",
        )

    # İsteğe bağlı group_column
    group_column = params.get("group_column")
    if group_column and group_column not in df.columns:
        raise HTTPException(
            status_code=400,
            detail=f"group_column '{group_column}' bulunamadı, mevcut sütunlar: {list(df.columns)}",
        )

    # İsteğe bağlı tarih filtreleme
    date_column = params.get("date_column")
    start_date = params.get("start_date")
    end_date = params.get("end_date")
    if date_column:
        if date_column not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"date_column '{date_column}' bulunamadı, mevcut sütunlar: {list(df.columns)}",
            )
        df[date_column] = pd.to_datetime(df[date_column], dayfirst=True, errors="coerce")
        if start_date:
            try:
                start_dt = pd.to_datetime(start_date, dayfirst=True)
                df = df[df[date_column] >= start_dt]
            except Exception:
                raise HTTPException(status_code=400, detail="start_date geçerli tarih değil")
        if end_date:
            try:
                end_dt = pd.to_datetime(end_date, dayfirst=True)
                df = df[df[date_column] <= end_dt]
            except Exception:
                raise HTTPException(status_code=400, detail="end_date geçerli tarih değil")

    # value_column sayısal hale getir
    df[value_column] = pd.to_numeric(df[value_column], errors="coerce")
    df = df.dropna(subset=[value_column])
    if df.empty:
        raise HTTPException(status_code=400, detail="value_column'da geçerli veri yok")

    # Eğer group_column varsa grupla ve ortalama al
    if group_column:
        summary_df = df.groupby(group_column, dropna=False)[value_column].mean().reset_index()
        summary_df = summary_df.sort_values(by=value_column)
    else:
        summary_df = df[[value_column]].copy()
        summary_df = summary_df.sort_values(by=value_column).reset_index(drop=True)

    # Excel dosyası oluştur
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Result"

    # DataFrame'i Excel'e yaz
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)

    # Renk skalası kuralı oluştur
    min_color = "FF63BE7B"  # Açık yeşil
    mid_color = "FFFFFF99"  # Açık sarı
    max_color = "FFFF5F5F"  # Açık kırmızı

    # value_column sütun indeksi (1-based)
    val_col_idx = summary_df.columns.get_loc(value_column) + 1
    first_data_row = 2
    last_data_row = first_data_row + len(summary_df) - 1
    cell_range = f"{ws.cell(row=first_data_row, column=val_col_idx).column_letter}{first_data_row}:{ws.cell(row=last_data_row, column=val_col_idx).column_letter}{last_data_row}"

    color_scale_rule = ColorScaleRule(
        start_type="min",
        start_color=min_color,
        mid_type="percentile",
        mid_value=50,
        mid_color=mid_color,
        end_type="max",
        end_color=max_color,
    )
    ws.conditional_formatting.add(cell_range, color_scale_rule)

    wb.save(output)
    output.seek(0)

    summary = {
        "rows": len(summary_df),
        "grouped_by": group_column if group_column else None,
        "value_column": value_column,
        "date_filtered": bool(date_column and (start_date or end_date)),
    }
    
    # Python kod özeti
    technical_details = {
        "scenario": "color_scale_by_value",
        "parameters": {"value_column": value_column, "group_column": group_column},
        "stats": {"rows": len(summary_df)},
        "python_code": f"""```python
import pandas as pd

df = pd.read_excel('dosya.xlsx')

# Renk skalası (Excel'de koşullu biçimlendirme ile yapılır)
# Veriler sıralanır, düşük=yeşil, orta=sarı, yüksek=kırmızı
sorted_df = df.sort_values('{value_column}')

# Excel'de ColorScale kuralı uygulanır
sorted_df.to_excel('renk_skalali.xlsx', index=False)
```"""
    }

    return {
        "summary": summary,
        "technical_details": technical_details,
        "df_out": summary_df,
        "excel_bytes": output,
        "excel_filename": "color_scale_result.xlsx",
    }