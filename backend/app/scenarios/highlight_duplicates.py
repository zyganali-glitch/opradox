from io import BytesIO
from typing import Any, Dict
import pandas as pd
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def run(df: pd.DataFrame, params: Dict[str, Any]) -> Dict[str, Any]:
    # Beklenen parametreler: columns (list of str) - kontrol edilecek sütunlar
    columns = params.get("columns")
    if not columns or not isinstance(columns, list):
        raise HTTPException(status_code=400, detail="Eksik veya hatalı 'columns' parametresi. Liste olmalı.")
    missing_cols = [col for col in columns if col not in df.columns]
    if missing_cols:
        raise HTTPException(status_code=400, detail=f"DataFrame'de bulunmayan sütunlar: {missing_cols}")

    # Sadece seçilen sütunları içeren alt df
    sub_df = df[columns].copy()

    # Duplicate kontrolü: her satırda seçilen sütunlardaki değerlerin tekrar edenlerini bul
    # Burada sütun bazında duplicate renklendirme yapılacak
    # Her sütun için duplicate olan hücreler True olacak şekilde mask oluştur
    duplicate_masks = {}
    for col in columns:
        # NaN'ler duplicate sayılmaz, onları False yap
        col_series = sub_df[col]
        mask = col_series.duplicated(keep=False) & col_series.notna()
        duplicate_masks[col] = mask

    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Result"

    # DataFrame'i Excel'e yaz
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)

    # Renklendirme için fill objesi
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Sarı

    # Sütunların Excel'deki indeksleri (1-based)
    col_idx_map = {col: idx + 1 for idx, col in enumerate(df.columns)}

    # Duplicate hücreleri renklendir
    for col in columns:
        mask = duplicate_masks[col]
        excel_col_idx = col_idx_map[col]
        for df_row_idx, is_dup in mask.items():
            if is_dup:
                excel_row_idx = df_row_idx + 2  # header + 1-based index
                ws.cell(row=excel_row_idx, column=excel_col_idx).fill = fill

    # BytesIO'ya kaydet
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    summary = {
        "checked_columns": columns,
        "total_rows": len(df),
        "duplicates_found": int(sum(mask.sum() for mask in duplicate_masks.values()))
    }
    
    # Python kod özeti
    cols_str = ', '.join([f"'{c}'" for c in columns])
    technical_details = {
        "scenario": "highlight_duplicates",
        "parameters": {"columns": columns},
        "stats": {"duplicates_found": summary["duplicates_found"]},
        "python_code": f"""```python
import pandas as pd

df = pd.read_excel('dosya.xlsx')

# Tekrarlayan değerleri bul
columns = [{cols_str}]
for col in columns:
    df[f'{{col}}_is_dup'] = df[col].duplicated(keep=False)

# Sonuç: {summary["duplicates_found"]} tekrarlayan hücre
df.to_excel('tekrarlar_isaretli.xlsx', index=False)
```"""
    }

    return {
        "summary": summary,
        "technical_details": technical_details,
        "df_out": df,
        "excel_bytes": excel_bytes,
        "excel_filename": "highlighted_duplicates.xlsx"
    }