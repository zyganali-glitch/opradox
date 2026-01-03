from io import BytesIO
from typing import Any, Dict
import pandas as pd
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def run(df: pd.DataFrame, params: Dict[str, Any]) -> Dict[str, Any]:
    # Parametre kontrolü
    value_column = params.get("value_column")
    n = params.get("n", 5)  # default=5
    mode = params.get("mode", "top").lower()  # default='top'
    group_column = params.get("group_column", None)

    # value_column boşsa ilk numeric auto-seç
    if not value_column:
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        if numeric_cols:
            value_column = numeric_cols[0]
        else:
            raise HTTPException(status_code=400, detail="Sayısal sütun bulunamadı. Lütfen value_column belirtin.")

    if value_column not in df.columns:
        raise HTTPException(
            status_code=400,
            detail=f"'{value_column}' sütunu bulunamadı, mevcut sütunlar: {list(df.columns)[:10]}",
        )

    if group_column is not None and group_column not in df.columns:
        raise HTTPException(
            status_code=400,
            detail=f"'{group_column}' sütunu bulunamadı, mevcut sütunlar: {list(df.columns)[:10]}",
        )

    try:
        n = int(n)
        if n <= 0:
            raise ValueError()
    except Exception:
        raise HTTPException(status_code=400, detail="'n' pozitif tam sayı olmalı")

    if mode not in ("top", "bottom"):
        raise HTTPException(status_code=400, detail="'mode' parametresi 'top' veya 'bottom' olmalı")

    try:
        n = int(n)
        if n <= 0:
            raise ValueError()
    except Exception:
        raise HTTPException(status_code=400, detail="'n' pozitif tam sayı olmalı")

    if mode not in ("top", "bottom"):
        raise HTTPException(status_code=400, detail="'mode' parametresi 'top' veya 'bottom' olmalı")

    # Değer sütununu sayısala çevir
    df[value_column] = pd.to_numeric(df[value_column], errors="coerce")

    # NaN değerleri hariç tut
    df_valid = df.dropna(subset=[value_column]).copy()

    # Renklendirme için renk seçimi
    fill_top = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Sarı
    fill_bottom = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # Mavi

    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Result"

    # Eğer group_column varsa gruplara göre işlem yap
    if group_column:
        # Gruplara göre top/bottom n belirle
        def highlight_group(group):
            if mode == "top":
                thresh = group[value_column].nlargest(n).min() if len(group) >= n else group[value_column].min() - 1
                group["highlight"] = group[value_column] >= thresh
            else:
                thresh = group[value_column].nsmallest(n).max() if len(group) >= n else group[value_column].max() + 1
                group["highlight"] = group[value_column] <= thresh
            return group

        df_valid = df_valid.groupby(group_column, group_keys=False).apply(highlight_group)
    else:
        if mode == "top":
            thresh = df_valid[value_column].nlargest(n).min() if len(df_valid) >= n else df_valid[value_column].min() - 1
            df_valid["highlight"] = df_valid[value_column] >= thresh
        else:
            thresh = df_valid[value_column].nsmallest(n).max() if len(df_valid) >= n else df_valid[value_column].max() + 1
            df_valid["highlight"] = df_valid[value_column] <= thresh

    # Orijinal df'ye highlight sütunu ekle (False olanlar da dahil)
    df["highlight"] = False
    df.loc[df_valid.index, "highlight"] = df_valid["highlight"]

    # Excel'e yaz
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            continue
        if df.loc[r_idx - 2, "highlight"]:
            fill = fill_top if mode == "top" else fill_bottom
            for c_idx in range(1, len(row) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = fill

    # Dosyayı BytesIO'ya kaydet
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    summary = {
        "mode": mode,
        "value_column": value_column,
        "n": n,
        "group_column": group_column,
        "highlighted_count": int(df["highlight"].sum()),
        "total_rows": len(df),
    }
    
    # Python kod özeti
    technical_details = {
        "scenario": "highlight_top_bottom_n",
        "parameters": {"value_column": value_column, "n": n, "mode": mode, "group_column": group_column},
        "stats": {"highlighted_count": int(df["highlight"].sum()), "total_rows": len(df)},
        "python_code": f"""```python
import pandas as pd

df = pd.read_excel('dosya.xlsx')

# {'En yüksek' if mode == 'top' else 'En düşük'} {n} değeri vurgula
df = df.sort_values('{value_column}', ascending={'True' if mode == 'bottom' else 'False'})
df['highlight'] = df.index.isin(df.head({n}).index)

# Sonuç: {int(df["highlight"].sum())} satır vurgulandı
df.to_excel('vurgulandi.xlsx', index=False)
```"""
    }

    return {
        "summary": summary,
        "technical_details": technical_details,
        "df_out": df,
        "excel_bytes": excel_bytes,
        "excel_filename": f"highlight_{mode}_{n}.xlsx",
    }